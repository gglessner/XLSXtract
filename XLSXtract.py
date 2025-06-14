#!/usr/bin/env python3

"""
XLSXtract - Extract text from Excel files for password generation
Copyright (C) 2024 Garland Glessner <gglessner@gmail.com>

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.

Author: Garland Glessner
Email: gglessner@gmail.com
"""

import argparse
import os
import sys
import warnings
import shutil
from pathlib import Path
from openpyxl import load_workbook
from typing import Set, Generator, List, Tuple
import time
import re

# Suppress openpyxl data validation warning
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def is_complex_password(word: str) -> bool:
    """Check if a word meets password complexity requirements."""
    has_upper = bool(re.search(r'[A-Z]', word))
    has_lower = bool(re.search(r'[a-z]', word))
    has_digit = bool(re.search(r'\d', word))
    # Common password special characters: !@#$%^&*()_+-=[]{}|;:,.<>?/~`
    has_special = bool(re.search(r'[!@#$%^&*()_+\-=\[\]{}|;:,.<>?/~`]', word))
    return has_upper and has_lower and has_digit and has_special

def get_terminal_width() -> int:
    """Get the width of the terminal window."""
    try:
        return shutil.get_terminal_size().columns
    except:
        return 80  # Default fallback width

def truncate_for_display(text: str, max_length: int) -> str:
    """Truncate text to fit terminal width, accounting for the 'Extracting: ' prefix."""
    prefix = "Extracting: "
    max_display_length = get_terminal_width() - len(prefix) - 3  # -3 for safety margin
    if len(text) > max_display_length:
        return text[:max_display_length] + "..."
    return text

def find_xlsx_files(directory: str, filename_pattern: str = None) -> Generator[Path, None, None]:
    """Recursively find all .xlsx files in the given directory.
    If filename_pattern is provided, only return files matching that name."""
    directory_path = Path(directory)
    if filename_pattern:
        # If filename_pattern doesn't end with .xlsx, add it
        if not filename_pattern.lower().endswith('.xlsx'):
            filename_pattern += '.xlsx'
        # Return only files matching the exact name
        return (f for f in directory_path.rglob("*.xlsx") if f.name.lower() == filename_pattern.lower())
    return directory_path.rglob("*.xlsx")

def print_progress(word: str, word_count: int):
    """Print progress with word truncation based on terminal width."""
    terminal_width = get_terminal_width()
    # Account for the "Extracting: " prefix and some padding
    max_word_width = terminal_width - 20
    truncated_word = truncate_for_display(word, max_word_width)
    # Clear the line and print with carriage return
    print(f"\rExtracting: {truncated_word} (Found {word_count} words)", end='', flush=True)

def extract_text_from_xlsx(xlsx_path: Path, split_chars: str, max_length: int, show_progress: bool = False, check_complexity: bool = False) -> Tuple[Set[str], int]:
    """Extract text values from an XLSX file."""
    text_values = set()
    word_count = 0
    skipped_words = 0
    
    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        
        for sheet in workbook:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # Convert to string and clean
                        text = str(cell.value).strip()
                        if not text:
                            continue
                            
                        # Split into words using specified delimiters
                        words_to_process = []
                        if split_chars:
                            # Create a regex pattern that matches any of the split characters
                            pattern = f"[{re.escape(split_chars)}]+"
                            words_to_process = re.split(pattern, text)
                        else:
                            words_to_process = [text]
                        
                        for word in words_to_process:
                            word = word.strip()
                            if not word or len(word) > max_length:
                                skipped_words += 1
                                continue
                                
                            # Clean the word
                            word = ''.join(c for c in word if c.isprintable() and not c.isspace())
                            
                            # Skip if complexity check fails
                            if check_complexity and not is_complex_password(word):
                                skipped_words += 1
                                continue
                                
                            if word and word not in text_values:
                                text_values.add(word)
                                word_count += 1
                                if show_progress:
                                    print(f"\033[K", end='')  # Clear the line
                                    print(f"\rExtracting: {truncate_for_display(word, get_terminal_width())}", end='', flush=True)
        
        if show_progress:
            print()  # New line after done with word display
        
        workbook.close()
        return text_values, word_count, skipped_words
        
    except Exception as e:
        print(f"\nError processing {xlsx_path}: {str(e)}")
        return set(), 0, 0

def process_xlsx_file(xlsx_path: Path, split_chars: str, show_progress: bool, max_length: int, check_complexity: bool) -> Tuple[Set[str], int]:
    """Process a single XLSX file and return the extracted text values."""
    print(f"Processing: {xlsx_path}")
    
    # Extract text from the Excel file
    text_values, word_count, skipped_words = extract_text_from_xlsx(xlsx_path, split_chars, max_length, show_progress, check_complexity)
    
    if show_progress:
        print()  # New line after done with word display
    else:
        print(f"Found {word_count} words")
        if check_complexity:
            print(f"Skipped {skipped_words} words that didn't meet complexity requirements")
    
    return text_values, skipped_words

def main():
    parser = argparse.ArgumentParser(description='Extract text from XLSX files for password generation.')
    parser.add_argument('-d', '--directory', required=True, help='Directory to scan for XLSX files')
    parser.add_argument('-o', '--output', default='passwords.txt', help='Output file for extracted text (default: passwords.txt)')
    parser.add_argument('-s', '--split-chars', default='', help='Characters to split words on (default: no splitting, process whole cell contents. Example: " ;:\'()<>\\"[]" for space, semicolon, colon, quotes, brackets, etc.)')
    parser.add_argument('-p', '--progress', action='store_true', help='Show real-time progress of each word being extracted (slower)')
    parser.add_argument('-l', '--max-length', type=int, default=32, help='Maximum length of words to extract (default: 32)')
    parser.add_argument('-f', '--filename', help='Only process files with this exact name (e.g., "Config.xlsx")')
    parser.add_argument('-c', '--complexity', action='store_true', help='Only extract words that meet password complexity requirements (uppercase, lowercase, number, and special character)')
    
    args = parser.parse_args()
    
    # Ensure the directory exists
    if not os.path.isdir(args.directory):
        print(f"Error: Directory '{args.directory}' does not exist")
        return
    
    # Statistics tracking
    total_files = 0
    total_words = 0
    total_skipped = 0
    all_words = set()
    
    # First pass: collect all words
    print("First pass: Collecting all words...")
    print(f"Maximum word length: {args.max_length} characters")
    if args.split_chars:
        print(f"Split characters: {args.split_chars}")
    if args.filename:
        print(f"Only processing files named: {args.filename}")
    if args.progress:
        print("Showing real-time word extraction (this will be slower)")
    if args.complexity:
        print("Checking password complexity (requires uppercase, lowercase, number, and special character)")
    xlsx_files = list(find_xlsx_files(args.directory, args.filename))
    
    if not xlsx_files:
        if args.filename:
            print(f"No files named '{args.filename}' found in {args.directory}")
        else:
            print(f"No XLSX files found in {args.directory}")
        return
    
    print(f"Found {len(xlsx_files)} XLSX files")
    
    # Process each file and collect words
    for xlsx_path in xlsx_files:
        words, skipped = process_xlsx_file(xlsx_path, args.split_chars, args.progress, args.max_length, args.complexity)
        all_words.update(words)
        total_files += 1
        total_words += len(words)
        total_skipped += skipped
    
    # Write final sorted and unique list
    print("\nWriting final sorted and unique word list...")
    with open(args.output, 'w', encoding='utf-8') as output_file:
        for word in sorted(all_words):
            output_file.write(f"{word}\n")
    
    # Print statistics
    print("\nProcessing complete!")
    print(f"Statistics:")
    print(f"- Files processed: {total_files}")
    print(f"- Total words found: {total_words}")
    print(f"- Unique words written: {len(all_words)}")
    print(f"- Maximum word length: {args.max_length} characters")
    if args.split_chars:
        print(f"- Split characters: {args.split_chars}")
    if args.filename:
        print(f"- Filename filter: {args.filename}")
    if args.complexity:
        print(f"- Words skipped (complexity): {total_skipped}")
    print(f"- Results written to: {args.output}")
    print()  # Extra blank line at end

if __name__ == "__main__":
    main() 